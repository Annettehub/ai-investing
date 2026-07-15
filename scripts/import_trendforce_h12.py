#!/usr/bin/env python3
"""Import TrendForce/DRAMeXchange metrics for the H1.2 dashboard.

CSV works without dependencies. XLSX support uses openpyxl when available.
The source workbook is never modified.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Iterable


ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "04-output" / "data" / "H1.2-trendforce.json"
EXPECTED = {"metric_id", "as_of", "value", "unit", "period", "source_note"}


def rows_from_csv(path: Path) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        yield from csv.DictReader(file)


def rows_from_xlsx(path: Path, sheet: str | None) -> Iterable[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("导入XLSX需要先安装 openpyxl；也可以把工作表另存为CSV。") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet] if sheet else workbook.active
    values = worksheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(values)]
    for row in values:
        yield dict(zip(headers, row))


def normalize(row: dict[str, Any]) -> dict[str, Any] | None:
    metric_id = str(row.get("metric_id") or "").strip()
    value = row.get("value")
    if not metric_id or value in (None, ""):
        return None
    try:
        numeric_value = float(str(value).replace(",", "").replace("%", ""))
    except ValueError as exc:
        raise RuntimeError(f"{metric_id} 的 value 不是数字: {value}") from exc
    return {
        "metricId": metric_id,
        "asOf": str(row.get("as_of") or "").strip(),
        "value": numeric_value,
        "unit": str(row.get("unit") or "").strip(),
        "period": str(row.get("period") or "").strip(),
        "sourceNote": str(row.get("source_note") or "").strip(),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Import TrendForce metrics for H1.2")
    parser.add_argument("source", type=Path, help="CSV or XLSX file")
    parser.add_argument("--sheet", help="XLSX sheet name; defaults to the active sheet")
    args = parser.parse_args()
    source = args.source.resolve()
    if not source.exists():
        parser.error(f"文件不存在: {source}")
    if source.suffix.lower() == ".csv":
        rows = rows_from_csv(source)
    elif source.suffix.lower() in {".xlsx", ".xlsm"}:
        rows = rows_from_xlsx(source, args.sheet)
    else:
        parser.error("只支持 CSV、XLSX 或 XLSM")

    metrics = [item for row in rows if (item := normalize(row))]
    if not metrics:
        raise RuntimeError("没有找到可导入的数据。请参考 config/trendforce/H1.2-template.csv。")
    payload = {
        "sourceFile": source.name,
        "importedAt": datetime.now(UTC).isoformat(timespec="seconds"),
        "asOf": max((item["asOf"] for item in metrics), default=""),
        "metrics": metrics,
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT.open("w", encoding="utf-8", newline="\n") as file:
        file.write(json.dumps(payload, ensure_ascii=False, indent=2) + "\n")
    print(f"Imported {len(metrics)} metrics to {OUTPUT.relative_to(ROOT)}")
    print("Run: python scripts/update_h12_dashboard_data.py --offline")
    return 0


if __name__ == "__main__":
    sys.exit(main())
